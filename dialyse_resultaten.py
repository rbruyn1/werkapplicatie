"""
dialyse_resultaten.py
=====================
Playwright-module voor dialyse staalname-resultaten:
  1. WO OPZOEKEN  — zoek de PeopleSoft werkorder op basis van T-nummer en staalnamedatum
  2. WO AFWERKEN  — vul endotoxine, kiemgetal, commentaar in en zet status op UITGEV

Configuratie via config.json (zelfde als wo_aanmaken.py):
    ps_base_url / ps_psc_url / username / password

Aanroep vanuit app.py:
    from dialyse_resultaten import zoek_werkorders, verzend_resultaten
    resultaten = asyncio.run(zoek_werkorders(stalen, stap_log=fn))
    resultaten = asyncio.run(verzend_resultaten(stalen, stap_log=fn))

Staalformaat (één staal = dict):
    {
      "datum":       "18/05/2026",   # staaldatum dd/mm/yyyy
      "eenheid":     "NIER 2",
      "detail":      "toestel TD51",
      "maximo":      "T301943",      # T-nummer
      "kiemgetal":   "0",            # CFU/ml numeriek als string
      "endotoxine":  "<0,125",       # inclusief < of >
    }
"""

import asyncio
import json
import os
import re
from datetime import datetime, timedelta
from pathlib import Path

import pdfplumber

# ── Configuratie ──────────────────────────────────────────────────────────────

def _lees_config() -> dict:
    pad = Path(__file__).parent / "config.json"
    if pad.exists():
        with open(pad, encoding="utf-8") as f:
            return json.load(f)
    return {}

_cfg = _lees_config()

PS_BASE_URL = os.getenv("PS_BASE_URL", _cfg.get("ps_base_url",
    _cfg.get("ps_base_url_td",
    "https://peoplesoftlogistiek.uz.kuleuven.ac.be/psp/FS9PROD")))
PS_PSC_URL  = os.getenv("PS_PSC_URL",  _cfg.get("ps_psc_url",
    "https://peoplesoftlogistiek.uz.kuleuven.ac.be/psc/FS9PROD/"))
PS_USER     = os.getenv("PS_USER",     _cfg.get("username", ""))
from crypto_utils import get_ps_password
PS_PASS     = os.getenv("PS_PASS",     get_ps_password(_cfg))
PS_HEADLESS = os.getenv("PS_HEADLESS", "false").lower() == "true"

LOGIN_URL   = PS_BASE_URL + "?cmd=login"

WO_ZOEK_URL = (
    PS_PSC_URL
    + "EMPLOYEE/ERP/c/UZFM_MENU.UZFM_WO_ZK.GBL"
    "?FolderPath=PORTAL_ROOT_OBJECT.UZFM.UZFM_WERKORDER.UZFM_WO_ZK_GBL"
    "&IsFolder=false&PortalHostNode=ERP&NoCrumbs=yes&PortalKeyStruct=yes"
)

# "Zoek Object" scherm — gebruikt om via roepnaam (bv. G06) het T-nummer op te
# zoeken. Zelfde URL/velden als de bewezen implementatie in
# wo_toestel_werkorder.py (zoek_tnummer_via_verkortingsnummer).
OBJECT_ZK_URL = (
    PS_PSC_URL.rstrip("/") + "/"
    + "EMPLOYEE/ERP/c/UZFM_MENU.UZFM_OBJECT_ZK.GBL?Folder=MYFAVORITES"
)


# ── Taak-ID mapping (conform VBA-macro) ──────────────────────────────────────

TAAK_ENDO_003  = "0000000222"   # endotoxine norm < 0,03 IU/ml  (toestellen 304 & 465)
TAAK_ENDO_025  = "0000000223"   # endotoxine norm < 0,25 IU/ml  (Genius / Water / thuisdialyse)
TAAK_MICRO_01  = "0000000224"   # microbiologie norm < 0,1 CFU/ml
TAAK_MICRO_100 = "0000000225"   # microbiologie norm < 100 CFU/ml
TAAK_KIEMGETAL = "0000000227"   # kiemgetal (vrije tekstwaarde)
TAAK_COMMENTAAR = "0000000226"  # commentaar labo


def _endo_taak_id(endotoxine: str) -> str:
    """
    Kies het juiste Taak-ID voor de endotoxine-norm.

    Het labo rapporteert de GEMETEN waarde (bv. <0,125 of <0,25),
    NIET de toegepaste norm. De norm voor dialyse/thuisdialyse is altijd
    <0,25 IU/ml (TAAK_ENDO_025 = 0000000223).

    Alleen voor toestellen 304 & 465 geldt de strengere norm <0,03 IU/ml
    (TAAK_ENDO_003 = 0000000222). Dat wordt herkend als de gemeten waarde
    zelf "0,03" of "0.03" bevat (bv. "<0,03").

    Voorbeelden labo-output -> Taak-ID:
        "<0,125"  ->  0000000223  (gemeten <0,125 is ruim binnen norm <0,25)
        "<0,25"   ->  0000000223
        ">0,25"   ->  0000000223  (buiten norm, maar zelfde taak-ID)
        "<0,03"   ->  0000000222  (strenge norm toestellen 304/465)
    """
    for s in ("0,03", "0.03"):
        if s in endotoxine:
            return TAAK_ENDO_003
    return TAAK_ENDO_025   # norm <0,25 voor dialyse / thuisdialyse


def _micro_taak_id(kiemgetal_norm: str) -> str:
    """Bepaal het juiste Taak-ID voor microbiologie (<100 CFU/ml voor dialyse/thuisdialyse)."""
    if "0,1" in kiemgetal_norm or "0.1" in kiemgetal_norm:
        return TAAK_MICRO_01
    return TAAK_MICRO_100   # default <100 CFU/ml voor dialyse/thuisdialyse


def _is_ok(waarde: str) -> bool:
    """
    < voor de waarde = binnen norm = True (OK).
    > voor de waarde = buiten norm = False.
    Labo geeft bv. "<0,125" (OK) of ">0,25" (buiten norm).
    """
    return "<" in waarde


# ── Login ────────────────────────────────────────────────────────────────────

async def login(page, log) -> bool:
    log(f"LOGIN: Navigeren naar {LOGIN_URL}")
    await page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=30_000)
    log(f"LOGIN: Pagina geladen — URL: {page.url[:80]}")
    await page.fill("#userid", PS_USER)
    await page.fill("#pwd", PS_PASS)
    await page.click("input[name='Submit']")
    log("LOGIN: Wachten op redirect...")
    try:
        await page.wait_for_url(
            lambda url: "EMPLOYEE" in url or "errorCode" in url,
            timeout=30_000
        )
    except Exception as e:
        log(f"LOGIN: Timeout bij redirect: {e}")
    if "errorCode" in page.url:
        log("LOGIN: ❌ Login mislukt (errorCode)")
        return False
    log("LOGIN: ✓ Login geslaagd")
    return True


# ── PDF-verwerking: Fresenius analyseverslag (Geniustoestel) ────────────────
#
# Verwacht formaat (zie voorbeeld 'Analyseverslag' Fresenius Medical Care):
#   "Datum staalname: 10.06.2026"
#   "Dialysaat: Geniustoestel nr. 6 /260610-078-04"
#   kolommen: Schimmels en gisten** | Totaal kiemgetal** | Endotoxines*
#   resultaatregel: "Genius 6   <0.1   <0.1   <0.0050"
#   checkbox: "☒ Er werden geen overschrijdingen vastgesteld t.o.v. de normen."

def _pdf_tekst(pdf_pad: str | Path) -> str:
    """Haal alle tekst op uit een PDF via pdfplumber."""
    tekst = ""
    try:
        with pdfplumber.open(str(pdf_pad)) as pdf:
            for page in pdf.pages:
                tekst += (page.extract_text() or "") + "\n"
    except Exception as e:
        tekst = f"[PDF LEESFOUT: {e}]"
    return tekst


def extraheer_staal_uit_pdf(pdf_pad: str | Path) -> dict:
    """
    Extraheer staalgegevens uit een Fresenius 'Analyseverslag' PDF
    (Geniustoestel dialysaat-resultaten).

    Geeft terug:
        {
          "ok": bool,
          "datum":        "10/06/2026",      # dd/mm/yyyy
          "toestel_nr":   "6",                # Geniustoestel-nummer
          "roepnaam":     "G06",              # afgeleide roepnaam voor object-zoek
          "eenheid":      "Genius 6",
          "schimmels":    "<0.1",
          "kiemgetal":    "<0.1",
          "endotoxine":   "<0.0050",
          "overschrijding": bool,
          "pdf_pad":      str(pdf_pad),
          "bestandsnaam": "....pdf",
          "fout":         str | None,
        }
    """
    pdf_pad = str(pdf_pad)
    info = {
        "ok": False, "datum": "", "toestel_nr": "", "roepnaam": "",
        "eenheid": "", "schimmels": "", "kiemgetal": "", "endotoxine": "",
        "overschrijding": False, "pdf_pad": pdf_pad,
        "bestandsnaam": Path(pdf_pad).name, "fout": None,
    }

    tekst = _pdf_tekst(pdf_pad)
    if tekst.startswith("[PDF LEESFOUT"):
        info["fout"] = tekst
        return info

    # ── Datum staalname (dd.mm.yyyy → dd/mm/yyyy) ───────────────────────────
    m = re.search(r"Datum staalname:\s*(\d{2})\.(\d{2})\.(\d{4})", tekst)
    if m:
        dd, mm, yyyy = m.groups()
        info["datum"] = f"{dd}/{mm}/{yyyy}"
    else:
        info["fout"] = "Datum staalname niet gevonden in PDF"

    # ── Geniustoestel nummer → roepnaam G0X ─────────────────────────────────
    m = re.search(r"Geniustoestel\s*nr\.?\s*(\d+)", tekst, re.IGNORECASE)
    if m:
        nr = m.group(1)
        info["toestel_nr"] = nr
        info["roepnaam"]   = f"G{int(nr):02d}"
        info["eenheid"]    = f"Genius {nr}"
    else:
        info["fout"] = (info["fout"] or "") + " | Geniustoestel-nummer niet gevonden in PDF"

    # ── Resultaatregel: "Genius [toestel] <nr>  <waarde>  <waarde>  [<waarde>]" ─
    # Sommige rapporten (bv. Geniustoestel 7) laten de kolom "Schimmels en
    # gisten" weg en rapporteren enkel Totaal kiemgetal + Endotoxines (2
    # waarden i.p.v. 3). Het label kan ook "Genius toestel <nr>" zijn i.p.v.
    # "Genius <nr>". We matchen daarom een variabel aantal waarden en
    # bepalen de kolom-mapping op basis van de headerregel.
    if info["toestel_nr"]:
        m = re.search(
            rf"Genius\s*(?:toestel\s*)?{re.escape(info['toestel_nr'])}\s+"
            r"((?:[<>]?\s*[\d.,]+|TNC)(?:\s+(?:[<>]?\s*[\d.,]+|TNC)){1,2})",
            tekst,
        )
        if m:
            waarden = re.findall(r"[<>]?\s*[\d.,]+|TNC", m.group(1))
            waarden = [w.replace(" ", "") for w in waarden]
            heeft_schimmels = "schimmels en gisten" in tekst.lower()

            if heeft_schimmels and len(waarden) == 3:
                info["schimmels"], info["kiemgetal"], info["endotoxine"] = waarden
            elif not heeft_schimmels and len(waarden) == 2:
                info["kiemgetal"], info["endotoxine"] = waarden
            elif len(waarden) == 3:
                # Onbekende header-combinatie, maar 3 waarden gevonden:
                # veronderstel schimmels/kiemgetal/endotoxine-volgorde.
                info["schimmels"], info["kiemgetal"], info["endotoxine"] = waarden
            elif len(waarden) == 2:
                info["kiemgetal"], info["endotoxine"] = waarden
            else:
                info["fout"] = (info["fout"] or "") + (
                    f" | Onverwacht aantal waarden ({len(waarden)}) in resultaatregel"
                )
        else:
            info["fout"] = (info["fout"] or "") + " | Resultaatregel niet gevonden in PDF"

    # ── Aangevinkt vakje opzoeken (checkbox-detectie via PDF-vormen) ────────
    # De vier vakjes in het analyseverslag zijn standaard-tekst; we bepalen
    # WELK vakje aangevinkt is via de tekenobjecten (rect + kruis-lijn),
    # niet via de (niet-extraheerbare) ☒/☐ glyphs.
    OPTIE_TEKSTEN = [
        "Er werden geen overschrijdingen vastgesteld t.o.v. de normen.",
        "Afwijkende resultaten: er werden één of meerdere overschrijdingen "
        "vastgesteld t.o.v. de normen.",
        "Indien analyse van permeaat/dialysaat: Er is een waarde boven de "
        "actielimiet (50% van de max. norm-waarde) voor de endotoxines gemeten.",
        "Indien analyse van permeaat/dialysaat: Er is een waarde boven de "
        "actielimiet (50% van de max. norm-waarde) voor het totaal kiemgetal gemeten.",
    ]
    checked_tekst = ""
    try:
        from collections import defaultdict
        with pdfplumber.open(pdf_pad) as pdf:
            page = pdf.pages[0]
            vierkantjes = [r for r in page.rects if 5 < r["width"] < 15 and 5 < r["height"] < 15]
            groepen = defaultdict(list)
            for r in vierkantjes:
                groepen[round(r["x0"], 1)].append(r)
            if groepen:
                optie_vakjes = sorted(max(groepen.values(), key=len), key=lambda r: -r["y0"])

                def _is_aangevinkt(r):
                    for l in page.lines:
                        if (l["x0"] >= r["x0"] - 2 and l["x1"] <= r["x1"] + 2 and
                                l["y0"] >= r["y0"] - 2 and l["y1"] <= r["y1"] + 2):
                            return True
                    return False

                for idx, vakje in enumerate(optie_vakjes):
                    if idx < len(OPTIE_TEKSTEN) and _is_aangevinkt(vakje):
                        checked_tekst = OPTIE_TEKSTEN[idx]
                        break
    except Exception as e:
        info["fout"] = (info["fout"] or "") + f" | Checkbox-detectie mislukt: {e}"

    info["checked_optie"]   = checked_tekst
    info["overschrijding"]  = bool(checked_tekst) and checked_tekst != OPTIE_TEKSTEN[0]

    info["ok"] = bool(info["datum"] and info["roepnaam"] and info["kiemgetal"] and info["endotoxine"])
    if not info["ok"] and not info["fout"]:
        info["fout"] = "Onvolledige gegevens uit PDF"

    return info


# ── T-nummer opzoeken via 'Zoek Object'-scherm (op roepnaam) ─────────────────

OBJ_ZOEK_BAS_DIENST  = "8"
OBJ_ZOEK_VAKGROEP_ID = "14550700"


async def zoek_tnummer_via_object(page, roepnaam: str, stap_log=None) -> dict:
    """
    Zoek het T-nummer op via het 'Zoek Object'-scherm, gefilterd op roepnaam
    (verkortingsnummer) + dienst + vakgroep.

    Zelfde bewezen flow/selectors als zoek_tnummer_via_verkortingsnummer() in
    wo_toestel_werkorder.py, maar hergebruikt hier de al ingelogde `page`
    zodat dit ingebed kan worden in verwerk_pdf_stalen() (één browsersessie
    voor object-zoek + WO-zoek per PDF).

    Geeft terug: {"ok": bool, "t_nummer": str, "omschrijving": str, "fout": str|None}
    """

    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] OBJ-ZOEK {roepnaam}: {msg}")
        if stap_log:
            stap_log(msg)

    resultaat = {"ok": False, "t_nummer": "", "omschrijving": "", "fout": None}

    try:
        log(f"Stap 1: Navigeren naar Zoek Object-scherm")
        await page.goto(OBJECT_ZK_URL, wait_until="domcontentloaded", timeout=30_000)
        await asyncio.sleep(2)

        label_veld    = "#UZFM_OBJ_ZK_WRK_UZ_OBJ_LABEL_ID"
        dienst_veld   = "#UZFM_OBJ_ZK_WRK_UZ_BAS_DIENST"
        vakgroep_veld = "#UZFM_OBJ_ZK_WRK_UZ_VAKGROEP_ID"
        zoek_knop     = "#UZFM_OBJ_ZK_WRK_UZ_ZOEK"

        log(f"Stap 2: Roepnaam ← '{roepnaam}'")
        await page.wait_for_selector(label_veld, state="visible", timeout=15_000)
        await page.fill(label_veld, roepnaam)

        log(f"Stap 3: Dienst ← '{OBJ_ZOEK_BAS_DIENST}'")
        try:
            await page.fill(dienst_veld, OBJ_ZOEK_BAS_DIENST)
        except Exception as e:
            log(f"Stap 3: ⚠ Dienst invullen mislukt: {e}")

        log(f"Stap 4: Vakgroep ← '{OBJ_ZOEK_VAKGROEP_ID}'")
        try:
            await page.fill(vakgroep_veld, OBJ_ZOEK_VAKGROEP_ID)
        except Exception as e:
            log(f"Stap 4: ⚠ Vakgroep invullen mislukt: {e}")

        log("Stap 5: [KLIK] Zoeken")
        await page.wait_for_selector(zoek_knop, state="visible", timeout=10_000)
        await page.click(zoek_knop)
        await page.wait_for_load_state("domcontentloaded", timeout=20_000)
        await asyncio.sleep(2)
        log("Stap 5: ✓ Zoekresultaten geladen")

        resultaat_sel = '[id="OBJECT$0"]'
        t_nr_tekst = ""
        try:
            await page.wait_for_selector(resultaat_sel, state="visible", timeout=10_000)
            t_nr_tekst = (await page.locator(resultaat_sel).inner_text()).strip()
            log(f"Stap 6: ✓ Eerste resultaat = '{t_nr_tekst}'")
        except Exception as e:
            log(f"Stap 6: ⚠ Geen resultaat voor '{roepnaam}': {e}")
            try:
                body = await page.locator("body").inner_text()
                log(f"Stap 6: Pagina-inhoud: {body[:200]}")
            except Exception:
                pass
            resultaat["fout"] = f"Geen object gevonden voor roepnaam '{roepnaam}'"
            return resultaat

        omschrijving = ""
        try:
            rij = page.locator(resultaat_sel).locator("xpath=ancestor::tr[1]")
            omschrijving = (await rij.inner_text(timeout=5_000)).strip()
            if t_nr_tekst and omschrijving.startswith(t_nr_tekst):
                omschrijving = omschrijving[len(t_nr_tekst):].strip(" \t-")
            log(f"Stap 7: Omschrijving = '{omschrijving[:80]}'")
        except Exception as e:
            log(f"Stap 7: Omschrijving niet leesbaar ({e})")

        resultaat["ok"]           = True
        resultaat["t_nummer"]     = t_nr_tekst
        resultaat["omschrijving"] = omschrijving
        return resultaat

    except Exception as exc:
        import traceback
        tb = traceback.format_exc()
        log(f"❌ Uitzondering: {exc}")
        for r in tb.splitlines():
            log(f"  TRACEBACK: {r}")
        resultaat["fout"] = str(exc)
        return resultaat


async def verwerk_pdf_stalen(pdf_paden: list[str], stap_log=None) -> list[dict]:
    """
    Volledige PDF-flow voor één of meerdere Fresenius analyseverslagen:
      1. PDF parsen → datum, roepnaam, resultaten
      2. T-nummer opzoeken via Zoek Object (roepnaam)
      3. WO opzoeken op T-nummer + staaldatum (hergebruik zoek_wo_voor_staal)

    Geeft een lijst stalen terug compatibel met het bestaande stalen-formaat
    (maximo, datum, eenheid, kiemgetal, endotoxine, ...), aangevuld met
    'pdf_pad' zodat afwerk_wo_voor_staal() de PDF als bijlage kan toevoegen.
    """
    from playwright.async_api import async_playwright

    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] PDF-VERWERK: {msg}")
        if stap_log:
            stap_log(msg)

    resultaten = []
    async with async_playwright() as pw:
        log(f"Browser opstarten (headless={PS_HEADLESS})...")
        browser = await pw.chromium.launch(headless=PS_HEADLESS)
        context = await browser.new_context(viewport={"width": 1280, "height": 900}, locale="nl-BE")
        page    = await context.new_page()
        log("Browser gereed ✓")
        try:
            if not await login(page, log):
                log("❌ Login mislukt")
                for p in pdf_paden:
                    resultaten.append({"pdf_pad": p, "ok": False, "fout": "Login mislukt"})
                return resultaten

            for idx, pdf_pad in enumerate(pdf_paden):
                bestandsnaam = Path(pdf_pad).name
                log(f"══ PDF {idx+1}/{len(pdf_paden)}: {bestandsnaam} ══")

                def s_log(msg, _b=bestandsnaam):
                    lbl = f"{_b}: {msg}"
                    ts  = datetime.now().strftime("%H:%M:%S")
                    print(f"[{ts}] {lbl}")
                    if stap_log:
                        stap_log(lbl)

                info = extraheer_staal_uit_pdf(pdf_pad)
                if not info["ok"]:
                    s_log(f"❌ PDF kon niet correct verwerkt worden: {info['fout']}")
                    resultaten.append({**info, "maximo": "", "wo_id": None, "status": None})
                    continue

                s_log(f"✓ Geparsed: datum={info['datum']}, roepnaam={info['roepnaam']}, "
                      f"kiemgetal={info['kiemgetal']}, endotoxine={info['endotoxine']}")

                obj_res = await zoek_tnummer_via_object(page, info["roepnaam"], stap_log=s_log)
                if not obj_res["ok"]:
                    s_log(f"❌ T-nummer niet gevonden: {obj_res['fout']}")
                    resultaten.append({**info, "maximo": "", "wo_id": None, "status": None,
                                        "fout": obj_res["fout"], "ok": False})
                    continue

                maximo = obj_res["t_nummer"]
                s_log(f"✓ T-nummer: {maximo} ({obj_res.get('omschrijving','')})")

                staal = {
                    "datum":       info["datum"],
                    "eenheid":     info["eenheid"],
                    "detail":      f"PDF: {bestandsnaam}",
                    "maximo":      maximo,
                    "kiemgetal":   info["kiemgetal"].lstrip("<>").strip(),
                    "endotoxine":  info["endotoxine"],
                    "commentaar":  info.get("checked_optie") or "*",
                    "pdf_pad":     str(pdf_pad),
                    "bestandsnaam": bestandsnaam,
                }

                wo_res = await zoek_wo_voor_staal(page, staal, stap_log=s_log)
                staal.update(wo_res)
                resultaten.append(staal)
                await asyncio.sleep(1)
        finally:
            log("Browser afsluiten...")
            await browser.close()
            log("Browser afgesloten ✓")

    return resultaten


# ── WO opzoeken voor één staal ───────────────────────────────────────────────

async def zoek_wo_voor_staal(page, staal: dict, stap_log=None) -> dict:
    """
    Zoek de werkorder op in PeopleSoft op basis van T-nummer en staalnamedatum.
    Geeft terug: {"wo_id": ..., "status": ..., "ok": bool, "fout": ...}
    """

    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] WO-zoek {staal.get('maximo','?')}: {msg}")
        if stap_log:
            stap_log(msg)

    maximo   = staal.get("maximo", "").strip()
    datum_str = staal.get("datum", "")  # dd/mm/yyyy

    # Bereken zoekvenster: 14 dagen voor t/m 14 dagen na de staaldatum
    try:
        staal_datum = datetime.strptime(datum_str, "%d/%m/%Y")
    except Exception:
        parsed = False
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%-m/%Y", "%-d/%-m/%Y", "%-d-%m-%Y", "%d-%m-%y"):
            try:
                staal_datum = datetime.strptime(datum_str, fmt)
                parsed = True
                break
            except Exception:
                continue
        if not parsed:
            # Laatste poging: dateutil indien beschikbaar
            try:
                from dateutil import parser as duparser
                staal_datum = duparser.parse(datum_str, dayfirst=True)
            except Exception:
                return {"ok": False, "wo_id": None, "status": None,
                        "fout": f"Ongeldige datum: {datum_str}"}

    start_range  = (staal_datum - timedelta(days=14)).strftime("%d/%m/%Y")
    end_range    = (staal_datum + timedelta(days=14)).strftime("%d/%m/%Y")

    log(f"Stap 1: Navigeren naar WO-zoekscherm")
    log(f"Stap 1: T-nummer={maximo}, zoekvenster={start_range} → {end_range}")

    try:
        await page.goto(WO_ZOEK_URL, wait_until="domcontentloaded", timeout=30_000)
        await page.wait_for_selector("#UZFM_WO_ZK_WRK_UZ_OMSCHR254", state="visible", timeout=20_000)
        log("Stap 1: ✓ Zoekscherm geladen")

        # Werktype = k (kwaliteit)
        log("Stap 2: Werktype 'k' ingeven")
        await page.click("#UZFM_WO_ZK_WRK_UZ_WO_TYPE")
        await page.keyboard.press("k")
        await page.keyboard.press("k")
        await page.click("#UZFM_WO_ZK_WRK_UZ_WO_TYPE")
        await asyncio.sleep(0.2)

        # Alleen open WO's uitzetten → ook afgesloten WO's doorzoeken
        await page.select_option("#UZFM_WO_ZK_WRK_UZ_WO_OPEN_ZK", "")
        await asyncio.sleep(0.2)

        # Begindatum zoekvenster
        log(f"Stap 3: Begindatum ← '{start_range}'")
        await page.click("#UZFM_WO_ZK_WRK_UZ_AANVRAAG_DT")
        await page.keyboard.press("Control+a")
        await asyncio.sleep(0.5)
        await page.type("#UZFM_WO_ZK_WRK_UZ_AANVRAAG_DT", start_range, delay=80)
        await page.keyboard.press("Tab")
        await asyncio.sleep(0.5)

        # Einddatum zoekvenster
        log(f"Stap 4: Einddatum ← '{end_range}'")
        await page.click("#UZFM_WO_ZK_WRK_END_DATE")
        await page.keyboard.press("Control+a")
        await asyncio.sleep(0.5)
        await page.type("#UZFM_WO_ZK_WRK_END_DATE", end_range, delay=80)
        await page.keyboard.press("Tab")
        await asyncio.sleep(0.5)

        # T-nummer — als laatste invullen (anders verdwijnen andere velden)
        log(f"Stap 5: T-nummer ← '{maximo}'")
        await page.click("#UZFM_WO_ZK_WRK_UZ_OBJ_ID")
        await page.keyboard.press("Control+a")
        await asyncio.sleep(0.2)
        await page.type("#UZFM_WO_ZK_WRK_UZ_OBJ_ID", maximo, delay=80)
        await page.keyboard.press("Tab")
        await asyncio.sleep(0.5)

        # Zoeken starten
        log("Stap 6: [KLIK] Zoeken")
        await page.click("#UZFM_WO_ZK_WRK_UZ_ZOEK")

        # Wachten op resultaat (max 30s / 10 pogingen à 3s)
        log("Stap 7: Wachten op zoekresultaten...")
        wo_id    = None
        wo_status = None
        for poging in range(10):
            await asyncio.sleep(3)
            aanwezig = await page.query_selector("#UZFM_WO_ZK_UZ_OMSCHR254\\$0")
            if aanwezig:
                try:
                    wo_id = (await page.locator("#UZ_WO_ID\\$0").inner_text(timeout=5_000)).strip()
                    wo_status = (await page.locator("#UZFM_WO_ZK_UZ_WO_STATUS\\$0").inner_text(timeout=5_000)).strip()
                    log(f"Stap 7: ✓ WO gevonden: {wo_id} — status: {wo_status}")
                except Exception as e:
                    log(f"Stap 7: Resultaat niet leesbaar: {e}")
                break
            log(f"Stap 7: Nog geen resultaat (poging {poging+1}/10)...")
        else:
            log("Stap 7: ℹ️ Geen werkorder gevonden (timeout)")
            return {"ok": True, "wo_id": "nvt", "status": "nvt",
                    "fout": None, "maximo": maximo, "datum": datum_str}

        if not wo_id:
            log("Stap 7: ⚠ WO-element gevonden maar nummer leeg")
            return {"ok": False, "wo_id": None, "status": None,
                    "fout": "WO-element gevonden maar nummer leeg",
                    "maximo": maximo, "datum": datum_str}

        return {"ok": True, "wo_id": wo_id, "status": wo_status,
                "fout": None, "maximo": maximo, "datum": datum_str}

    except Exception as exc:
        import traceback
        tb = traceback.format_exc()
        log(f"❌ Uitzondering: {exc}")
        for r in tb.splitlines():
            log(f"  TRACEBACK: {r}")
        return {"ok": False, "wo_id": None, "status": None,
                "fout": str(exc), "maximo": maximo, "datum": datum_str}


# ── WO afwerken voor één staal ────────────────────────────────────────────────

async def afwerk_wo_voor_staal(page, staal: dict, stap_log=None) -> dict:
    """
    Navigeer naar de werkorder, vul resultaten in en zet status op UITGEV.
    staal moet ook 'wo_id' bevatten (gezet door zoek_wo_voor_staal).
    """

    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] WO-afwerk {staal.get('maximo','?')} WO{staal.get('wo_id','?')}: {msg}")
        if stap_log:
            stap_log(msg)

    maximo      = staal.get("maximo", "").strip()
    wo_id       = staal.get("wo_id", "").strip().zfill(10)
    datum_str   = staal.get("datum", "")
    eenheid     = staal.get("eenheid", "")
    kiemgetal   = str(staal.get("kiemgetal", "0")).strip()
    endotoxine  = staal.get("endotoxine", "").strip()
    commentaar  = staal.get("commentaar", "*").strip() or "*"

    if not wo_id or wo_id == "0000000000" or wo_id.lower() == "nvt":
        log("⚠ Geen geldig WO-nummer — overslaan")
        return {"ok": False, "wo_id": wo_id, "status": None,
                "fout": "Geen geldig WO-nummer", "maximo": maximo}

    wo_url = (
        PS_PSC_URL
        + "EMPLOYEE/ERP/c/UZFM_MENU.UZFM_WERKORDER.GBL"
        f"?BUSINESS_UNIT=POUZL&UZ_WO_ID={wo_id}"
        "&PAGE=UZFM_WO_ALG&PortalContentProvider=ERP"
        "&PortalCRefLabel=Werkorder&PortalRegistryName=EMPLOYEE"
        "&PortalHostNode=ERP&NoCrumbs=yes&PortalKeyStruct=yes"
    )

    log(f"Stap 1: Navigeren naar WO {wo_id}")
    try:
        await page.goto(wo_url, wait_until="domcontentloaded", timeout=30_000)
        await page.wait_for_selector("#UZFM_WO_UZ_WO_ID", state="visible", timeout=20_000)

        # Valideer WO-nummer op pagina
        pagina_wo = (await page.locator("#UZFM_WO_UZ_WO_ID").inner_text(timeout=5_000)).strip()
        log(f"Stap 1: Pagina WO-nummer = '{pagina_wo}'")
        if wo_id not in pagina_wo and pagina_wo not in wo_id:
            log(f"Stap 1: ❌ Verkeerde WO geopend (verwacht {wo_id}, got {pagina_wo})")
            return {"ok": False, "wo_id": wo_id, "status": None,
                    "fout": f"Verkeerde WO: {pagina_wo}", "maximo": maximo}
        log("Stap 1: ✓ Correcte WO geopend")

        # Navigeer naar rapportage-tab (tab 2 = "#ICPanel2" of "#ICTAB_2")
        log("Stap 2: [KLIK] Rapportage-tab")
        for selector in ["#ICPanel2", "#ICTAB_2"]:
            try:
                btn = await page.query_selector(selector)
                if btn:
                    await page.click(selector)
                    break
            except Exception:
                pass
        await page.wait_for_selector("#UZFM_WO_WRK_UZ_WO_COMMENT", state="visible", timeout=20_000)
        log("Stap 2: ✓ Rapportage-tab geladen")

        # Status → UITGEV
        log("Stap 3: Status instellen op UITGEV (U)")
        await page.select_option("#UZFM_WO_WRK_UZ_WO_STAT_WZ_TECH\\$70\\$", "UITGEV")
        await asyncio.sleep(0.5)

        # Controleer of taaklijn 0 al ingevuld is
        eerste_taak = ""
        try:
            eerste_taak = (await page.locator("#UZFM_WO_TAAK_UZ_TAAK_ID\\$0").inner_text(timeout=3_000)).strip()
        except Exception:
            pass

        if len(eerste_taak) >= 2:
            log(f"Stap 4: ℹ️ Taaklijn 0 al ingevuld ({eerste_taak}) — WO overslaan")
            return {"ok": True, "wo_id": wo_id, "status": "reeds_ingevuld",
                    "fout": None, "maximo": maximo,
                    "melding": "Reeds ingevuld, overgeslaan"}

        # ── Werkorder commentaar ──────────────────────────────────────────
        commentaar_wo = f"Resultaat-> Staaldatum: {datum_str}  Toestel: {eenheid} ({maximo})"
        log(f"Stap 4: Commentaar ← '{commentaar_wo}'")
        await page.fill("#UZFM_WO_WRK_UZ_WO_COMMENT", commentaar_wo)

        # ── Lijn 0 — Endotoxine ───────────────────────────────────────────
        endo_taak = _endo_taak_id(endotoxine)
        endo_ok   = _is_ok(endotoxine)
        log(f"Stap 5: Endotoxine taak={endo_taak}, waarde='{endotoxine}', OK={endo_ok}")

        await page.type("#UZFM_WO_TAAK_UZ_TAAK_ID\\$0", endo_taak, delay=80)
        await page.click("#UZFM_WO_WRK_UZ_WO_COMMENT")  # trigger PS validatie
        await asyncio.sleep(1)
        # Wacht op omschrijving zoals VBA (TAAKDVW$0 niet leeg)
        for _ in range(5):
            try:
                tekst = (await page.locator("#UZFM_WO_TAAKDVW_UZ_OMSCHR150\\$0").inner_text(timeout=2_000)).strip()
                if tekst:
                    log(f"Stap 5: Omschrijving endo = '{tekst}'")
                    break
            except Exception:
                pass
            await asyncio.sleep(1)

        if endo_ok:
            log("Stap 5: [KLIK] Endotoxine → Ja")
            await page.click("#UZFM_WO_WRK_UZ_WAARDE_VLAG_JA\\$0")
        else:
            log("Stap 5: [KLIK] Endotoxine → Nee")
            await page.click("#UZFM_WO_WRK_UZ_WAARDE_VLAG_NEE\\$0")
        await asyncio.sleep(1)

        # ── Lijn 1 — Microbiologie (ja/nee norm) ─────────────────────────
        log("Stap 6: Nieuwe taaklijn toevoegen voor microbiologie")
        await page.click("#UZFM_WO_TAAK\\$new\\$0\\$\\$0")
        await page.wait_for_selector("#UZ_CHANGED_BY\\$1", state="visible", timeout=15_000)

        # Kiemgetal norm: < 100 CFU/ml voor dialyse/thuisdialyse
        micro_taak = _micro_taak_id(kiemgetal)
        # Bepaal OK op basis van numerieke kiemgetal vs norm
        try:
            kgetal = float(kiemgetal.replace(",", "."))
            if TAAK_MICRO_100 == micro_taak:
                micro_ok = kgetal < 100
            else:
                micro_ok = kgetal < 0.1
        except Exception:
            micro_ok = "<" in kiemgetal

        log(f"Stap 6: Microbiologie taak={micro_taak}, kiemgetal={kiemgetal}, OK={micro_ok}")
        await page.type("#UZFM_WO_TAAK_UZ_TAAK_ID\\$1", micro_taak, delay=80)
        await page.click("#UZFM_WO_WRK_UZ_WO_COMMENT")
        await asyncio.sleep(1)
        # Wacht op omschrijving zoals VBA (TAAKDVW$1 niet leeg)
        for _ in range(5):
            try:
                tekst = (await page.locator("#UZFM_WO_TAAKDVW_UZ_OMSCHR150\\$1").inner_text(timeout=2_000)).strip()
                if tekst:
                    log(f"Stap 6: Omschrijving micro = '{tekst}'")
                    break
            except Exception:
                pass
            await asyncio.sleep(1)

        if micro_ok:
            log("Stap 6: [KLIK] Microbiologie → Ja")
            await page.click("#UZFM_WO_WRK_UZ_WAARDE_VLAG_JA\\$1")
        else:
            log("Stap 6: [KLIK] Microbiologie → Nee")
            await page.click("#UZFM_WO_WRK_UZ_WAARDE_VLAG_NEE\\$1")
        await asyncio.sleep(1)

        # ── Lijn 2 — Kiemgetal (vrije tekst) ─────────────────────────────
        log("Stap 7: Nieuwe taaklijn voor kiemgetal (tekst)")
        await page.click("#UZFM_WO_TAAK\\$new\\$1\\$\\$0")
        await page.wait_for_selector("#UZ_CHANGED_BY\\$2", state="visible", timeout=15_000)

        log(f"Stap 7: Kiemgetal taak={TAAK_KIEMGETAL}, waarde='{kiemgetal}'")
        await page.type("#UZFM_WO_TAAK_UZ_TAAK_ID\\$2", TAAK_KIEMGETAL, delay=80)
        await page.click("#UZFM_WO_WRK_UZ_WO_COMMENT")
        await asyncio.sleep(1)
        # Wacht op omschrijving zoals VBA (TAAKDVW$2 niet leeg)
        for _ in range(5):
            try:
                tekst = (await page.locator("#UZFM_WO_TAAKDVW_UZ_OMSCHR150\\$2").inner_text(timeout=2_000)).strip()
                if tekst:
                    log(f"Stap 7: Omschrijving kiemgetal = '{tekst}'")
                    break
            except Exception:
                pass
            await asyncio.sleep(1)
        await page.type("#UZFM_WO_WRK_UZ_WAARDE_TEKST\\$2", kiemgetal, delay=80)
        await asyncio.sleep(1)

        # ── Lijn 3 — Commentaar labo (optioneel) ─────────────────────────
        if commentaar and commentaar != "*":
            log(f"Stap 8: Commentaar labo ← '{commentaar}'")
            await page.click("#UZFM_WO_TAAK\\$new\\$2\\$\\$0")
            await page.wait_for_selector("#UZ_CHANGED_BY\\$3", state="visible", timeout=15_000)

            await page.type("#UZFM_WO_TAAK_UZ_TAAK_ID\\$3", TAAK_COMMENTAAR, delay=80)
            await page.click("#UZFM_WO_WRK_UZ_WO_COMMENT")
            await asyncio.sleep(1)
            for _ in range(5):
                try:
                    tekst = (await page.locator("#UZFM_WO_TAAKDVW_UZ_OMSCHR150\\$3").inner_text(timeout=2_000)).strip()
                    if tekst:
                        log(f"Stap 8: Omschrijving commentaar = '{tekst}'")
                        break
                except Exception:
                    pass
                await asyncio.sleep(1)
            await page.fill("#UZFM_WO_WRK_UZ_WAARDE_TEKST\\$3", commentaar)
            await asyncio.sleep(0.5)
        else:
            log("Stap 8: Geen commentaar (overgeslagen)")

        # ── PDF-bijlage toevoegen (indien afkomstig van PDF-staalverwerking) ──
        pdf_pad = staal.get("pdf_pad", "")
        if pdf_pad and Path(pdf_pad).exists():
            log(f"Stap 8b: PDF-bijlage toevoegen: {Path(pdf_pad).name}")
            try:
                await page.click("#ICTAB_4")
                await asyncio.sleep(2)
                await page.wait_for_selector("#UZ_ATTACH_BTN\\$0", state="visible", timeout=10_000)
                await page.click("#UZ_ATTACH_BTN\\$0")
                await asyncio.sleep(3)

                await page.wait_for_selector("#UZ_ATTACH_WRK_UZ_ATTACH_ADD", state="visible", timeout=10_000)
                await page.click("#UZ_ATTACH_WRK_UZ_ATTACH_ADD")
                await asyncio.sleep(2)

                upload_modal = next((f for f in page.frames if f.name == "ptModFrame_0"), None)
                if not upload_modal:
                    log("Stap 8b: ❌ Upload modal niet gevonden — bijlage overgeslagen")
                else:
                    await asyncio.sleep(1)
                    file_input = await upload_modal.query_selector("input[type='file']")
                    if not file_input:
                        log("Stap 8b: ❌ File input niet gevonden — bijlage overgeslagen")
                    else:
                        await file_input.set_input_files(pdf_pad)
                        await asyncio.sleep(1)
                        try:
                            await upload_modal.wait_for_selector(
                                "#Upload:not([disabled])", state="visible", timeout=10_000)
                            await upload_modal.click("#Upload")
                            await asyncio.sleep(3)
                        except Exception as e:
                            log(f"Stap 8b: ⚠ Upload-knop niet enabled: {e}")

                        for _ in range(30):
                            await asyncio.sleep(1)
                            try:
                                el = await page.query_selector("#UZ_ATTACHMENT_ATTACHUSERFILE")
                                if el and (await el.inner_text()).strip():
                                    break
                            except Exception:
                                pass

                        try:
                            await page.fill("#UZ_ATTACHMENT_DESCR50_MIXED", "Analyseverslag Genius")
                        except Exception:
                            pass

                        try:
                            await page.click('[id="#ICSave"]')
                            await asyncio.sleep(3)
                        except Exception as e:
                            log(f"Stap 8b: ⚠ OK-knop bijlage-modal mislukt: {e}")

                        log("Stap 8b: ✓ PDF-bijlage toegevoegd")
            except Exception as e:
                log(f"Stap 8b: ⚠ PDF-bijlage toevoegen mislukt (WO afwerken gaat door): {e}")

            # terug naar rapportage-tab voor verdere afwerking
            for selector in ["#ICPanel2", "#ICTAB_2"]:
                try:
                    btn = await page.query_selector(selector)
                    if btn:
                        await page.click(selector)
                        break
                except Exception:
                    pass
            await page.wait_for_selector("#UZFM_WO_WRK_UZ_WO_COMMENT", state="visible", timeout=20_000)

        # ── Opslaan ───────────────────────────────────────────────────────
        log("Stap 9: [KLIK] Opslaan '#ICSave'")
        await page.click('[id="#ICSave"]')
        await asyncio.sleep(3)

        # ── Verificatie: heropen WO en controleer ─────────────────────────
        log("Stap 10: Heropen WO voor verificatie...")
        await page.goto(wo_url, wait_until="domcontentloaded", timeout=30_000)
        await page.wait_for_selector("#UZFM_WO_UZ_WO_ID", state="visible", timeout=20_000)

        for selector in ["#ICPanel2", "#ICTAB_2"]:
            try:
                btn = await page.query_selector(selector)
                if btn:
                    await page.click(selector)
                    break
            except Exception:
                pass
        await page.wait_for_selector("#UZFM_WO_WRK_UZ_WO_COMMENT", state="visible", timeout=20_000)

        taak0 = taak1 = taak2 = ""
        try:
            taak0 = (await page.locator("#UZFM_WO_TAAK_UZ_TAAK_ID\\$0").inner_text(timeout=3_000)).strip()
        except Exception:
            pass
        try:
            taak1 = (await page.locator("#UZFM_WO_TAAK_UZ_TAAK_ID\\$1").inner_text(timeout=3_000)).strip()
        except Exception:
            pass
        try:
            taak2 = (await page.locator("#UZFM_WO_TAAK_UZ_TAAK_ID\\$2").inner_text(timeout=3_000)).strip()
        except Exception:
            pass

        laatste_log = ""
        try:
            laatste_log = (await page.locator("#UZFM_WO_LOG_UZ_OMSCHR254\\$0").inner_text(timeout=3_000)).strip()
        except Exception:
            pass

        log(f"Stap 10: Verificatie — taak0={taak0}, taak1={taak1}, taak2={taak2}")
        log(f"Stap 10: Laatste historielijn = '{laatste_log}'")

        endo_taakids   = {TAAK_ENDO_003, TAAK_ENDO_025}
        micro_taakids  = {TAAK_MICRO_01, TAAK_MICRO_100}
        commentaar_verwacht = f"Resultaat-> Staaldatum: {datum_str}  Toestel: {eenheid} ({maximo})"

        # Eindstatus ophalen
        eindstatus = ""
        try:
            eindstatus = (await page.locator("#UZFM_WO_UZ_WO_STATUS").inner_text(timeout=5_000)).strip()
        except Exception:
            pass
        log(f"Stap 11: Eindstatus WO = '{eindstatus}'")

        # Commentaar vergelijking: normaliseer spaties
        commentaar_ok = laatste_log.split() == commentaar_verwacht.split()
        if not commentaar_ok:
            log(f"Stap 10: ⚠ Commentaar verschilt (verwacht='{commentaar_verwacht}', gevonden='{laatste_log}')")

        verzonden_ok = (
            taak0 in endo_taakids and
            taak1 in micro_taakids and
            taak2 == TAAK_KIEMGETAL and
            eindstatus == "UITGEV"
        )

        if verzonden_ok:
            log("Stap 10: ✓ Resultaten succesvol verzonden en geverifieerd")
            melding = "Resultaten succesvol verzonden"
        else:
            log("Stap 10: ⚠ Verificatie niet volledig geslaagd")
            melding = (
                f"taak0={taak0}, taak1={taak1}, taak2={taak2}, "
                f"status={eindstatus}, commentaar={'✓' if commentaar_ok else '✗'}"
            )

        return {
            "ok": verzonden_ok,
            "wo_id": wo_id,
            "status": eindstatus,
            "fout": None if verzonden_ok else melding,
            "maximo": maximo,
            "melding": melding,
        }

    except Exception as exc:
        import traceback
        tb = traceback.format_exc()
        log(f"❌ Uitzondering: {exc}")
        for r in tb.splitlines():
            log(f"  TRACEBACK: {r}")
        return {"ok": False, "wo_id": wo_id, "status": None,
                "fout": str(exc), "maximo": maximo}


# ── Hoofd entry-points ────────────────────────────────────────────────────────

async def zoek_werkorders(stalen: list[dict], stap_log=None) -> list[dict]:
    """
    Login eenmalig, zoek voor alle stalen de WO op.
    Geeft lijst terug met zelfde staal-dicts, uitgebreid met wo_id en status.
    """
    from playwright.async_api import async_playwright

    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] ZOEK: {msg}")
        if stap_log:
            stap_log(msg)

    resultaten = []

    async with async_playwright() as pw:
        log(f"Browser opstarten (headless={PS_HEADLESS})...")
        browser = await pw.chromium.launch(headless=PS_HEADLESS)
        context = await browser.new_context(viewport={"width": 1280, "height": 900}, locale="nl-BE")
        page    = await context.new_page()
        log("Browser gereed ✓")

        try:
            if not await login(page, log):
                log("❌ Login mislukt")
                for s in stalen:
                    resultaten.append({**s, "ok": False, "wo_id": None,
                                       "status": None, "fout": "Login mislukt"})
                return resultaten

            log(f"✓ Ingelogd — {len(stalen)} staal/stalen verwerken")
            for idx, staal in enumerate(stalen):
                log(f"══ Staal {idx+1}/{len(stalen)}: {staal.get('maximo','?')} ({staal.get('datum','?')}) ══")

                def s_log(msg, _s=staal):
                    lbl = f"{_s.get('maximo','?')} | {_s.get('datum','?')}: {msg}"
                    ts  = datetime.now().strftime("%H:%M:%S")
                    print(f"[{ts}] {lbl}")
                    if stap_log:
                        stap_log(lbl)

                try:
                    res = await zoek_wo_voor_staal(page, staal, stap_log=s_log)
                except Exception as exc:
                    import traceback
                    tb = traceback.format_exc()
                    log(f"❌ Onverwachte fout bij {staal.get('maximo','?')}: {exc}")
                    for r in tb.splitlines():
                        log(f"  {r}")
                    res = {"ok": False, "wo_id": None, "status": None, "fout": str(exc)}
                resultaten.append({**staal, **res})
                await asyncio.sleep(1)

        finally:
            log("Browser afsluiten...")
            await browser.close()
            log("Browser afgesloten ✓")

    return resultaten


async def verzend_resultaten(stalen: list[dict], stap_log=None) -> list[dict]:
    """
    Login eenmalig, werk voor alle stalen (met wo_id + status == INUITV) de WO af.
    """
    from playwright.async_api import async_playwright

    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] VERZEND: {msg}")
        if stap_log:
            stap_log(msg)

    resultaten = []

    async with async_playwright() as pw:
        log(f"Browser opstarten (headless={PS_HEADLESS})...")
        browser = await pw.chromium.launch(headless=PS_HEADLESS)
        context = await browser.new_context(viewport={"width": 1280, "height": 900}, locale="nl-BE")
        page    = await context.new_page()
        log("Browser gereed ✓")

        try:
            if not await login(page, log):
                log("❌ Login mislukt")
                for s in stalen:
                    resultaten.append({**s, "ok": False, "fout": "Login mislukt"})
                return resultaten

            log(f"✓ Ingelogd — {len(stalen)} staal/stalen ontvangen")
            for s in stalen:
                log(f"  → {s.get('maximo','?')} | wo_id={s.get('wo_id','?')} | status={s.get('status','?')}")

            te_verzenden = [s for s in stalen if s.get("wo_id") and s.get("wo_id") != "nvt"]
            overgeslagen = [s for s in stalen if s not in te_verzenden]

            log(f"  {len(te_verzenden)} te verzenden, {len(overgeslagen)} overgeslagen")
            for s in overgeslagen:
                reden = "Geen WO" if not s.get("wo_id") else f"Status = {s.get('status','?')}"
                log(f"  ↷ Overgeslagen: {s.get('maximo','?')} — {reden}")
                resultaten.append({**s, "ok": False, "melding": reden, "fout": reden})

            for idx, staal in enumerate(te_verzenden):
                log(f"══ Staal {idx+1}/{len(te_verzenden)}: {staal.get('maximo','?')} WO={staal.get('wo_id','?')} ══")

                def s_log(msg, _s=staal):
                    lbl = f"{_s.get('maximo','?')} WO{_s.get('wo_id','?')}: {msg}"
                    ts  = datetime.now().strftime("%H:%M:%S")
                    print(f"[{ts}] {lbl}")
                    if stap_log:
                        stap_log(lbl)

                res = await afwerk_wo_voor_staal(page, staal, stap_log=s_log)
                resultaten.append({**staal, **res})
                await asyncio.sleep(1)

        finally:
            log("Browser afsluiten...")
            await browser.close()
            log("Browser afgesloten ✓")

    return resultaten


# ── Excel import helper ───────────────────────────────────────────────────────

def importeer_excel(pad: str) -> list[dict]:
    """
    Lees het labo-Excel (Onderzoeksverslag waterstalen) en geef lijst van staal-dicts.
    Gebruikt openpyxl (geen pandas vereist).
    Zoekt automatisch de headerrij met 'Datum staalname'.
    """
    import openpyxl
    from datetime import datetime

    wb = openpyxl.load_workbook(pad, data_only=True)

    # Zoek sheet
    sheet = wb.worksheets[0]
    for ws in wb.worksheets:
        if "water" in ws.title.lower() or "report" in ws.title.lower():
            sheet = ws
            break

    # Zoek headerrij
    header_row = None
    header_idx = {}
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for j, cell in enumerate(row):
            if isinstance(cell, str) and "datum staalname" in cell.lower():
                header_row = i
                # Sla alle kolomindices op
                for k, h in enumerate(row):
                    if h is not None:
                        header_idx[str(h).strip().lower()] = k
                break
        if header_row:
            break

    if header_row is None:
        raise ValueError("Headerrij 'Datum staalname' niet gevonden in Excel")

    def vind_col(keywords):
        for key, idx in header_idx.items():
            if any(k.lower() in key for k in keywords):
                return idx
        return None

    ci_datum   = vind_col(["datum staalname", "datum"])
    ci_eenheid = vind_col(["eenheid"])
    ci_detail  = vind_col(["detail"])
    ci_maximo  = vind_col(["maximo"])
    ci_kiem    = vind_col(["kiemgetal", "cfu"])
    ci_endo    = vind_col(["endotoxine", "iu/ml"])

    stalen = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        maximo = str(row[ci_maximo]).strip() if ci_maximo is not None and row[ci_maximo] is not None else ""
        if not maximo or maximo.lower() in ("nan", "none", ""):
            continue

        # Datum formatteren
        datum_raw = row[ci_datum] if ci_datum is not None else None
        try:
            if isinstance(datum_raw, datetime):
                datum_str = datum_raw.strftime("%d/%m/%Y")
            elif isinstance(datum_raw, (int, float)):
                from datetime import timedelta
                d = datetime(1899, 12, 30) + timedelta(days=int(datum_raw))
                datum_str = d.strftime("%d/%m/%Y")
            elif isinstance(datum_raw, str) and datum_raw.strip():
                datum_str = datum_raw.strip()
            else:
                datum_str = ""
        except Exception:
            datum_str = str(datum_raw).strip() if datum_raw else ""

        def cel(idx):
            if idx is None or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        stalen.append({
            "datum":      datum_str,
            "eenheid":    cel(ci_eenheid),
            "detail":     cel(ci_detail),
            "maximo":     maximo,
            "kiemgetal":  cel(ci_kiem) or "0",
            "endotoxine": cel(ci_endo) or "<0,125",
            "commentaar": "*",
            "wo_id":      None,
            "status":     None,
        })

    return stalen

# ═══════════════════════════════════════════════════════════════════════════════
# RO RINGLEIDING — taak-IDs, importfunctie en afwerkfunctie
# ═══════════════════════════════════════════════════════════════════════════════

# Taak-IDs voor RO ringleiding (conform opgave)
TAAK_RO_ENDO_BEGIN  = "0000001949"   # Endotoxine <0,25 IU/ml — begin ringleiding
TAAK_RO_ENDO_EINDE  = "0000001950"   # Endotoxine <0,25 IU/ml — einde ringleiding
TAAK_RO_MICRO_BEGIN = "0000001951"   # Microbiologisch < 100 CFU/ml — begin ringleiding
TAAK_RO_MICRO_EINDE = "0000001952"   # Microbiologisch < 100 CFU/ml — einde ringleiding
TAAK_RO_KIEM_BEGIN  = "0000001953"   # Kiemgetal tekstwaarde — begin ringleiding
TAAK_RO_KIEM_EINDE  = "0000001954"   # Kiemgetal tekstwaarde — einde ringleiding


def _ro_kiemgetal_ok(waarde: str) -> bool:
    """
    Bepaal of het kiemgetal binnen de norm valt (< 100 CFU/ml).
    LM = minstens 300 CFU/l geteld = 0,3 CFU/ml  → buiten norm (>0,1 maar <100)
         maar labo gebruikt deze melding als indicatief te hoog -> Nee
    M  = >600 CFU/l = 0,6 CFU/ml  → Nee
    Numeriek: OK als < 100.
    """
    v = str(waarde).strip().upper()
    if v in ("LM", "M"):
        return False      # twijfelachtig / te hoog
    if "<" in v:
        return True
    if ">" in v:
        return False
    try:
        return float(v.replace(",", ".")) < 100
    except Exception:
        return True       # onbekend formaat, voorzichtig True


def _ro_endo_ok(waarde: str) -> bool:
    """< in de waarde = binnen norm."""
    return "<" in str(waarde)


def importeer_excel_ro(pad: str) -> list[dict]:
    """
    Lees het RO-Excel (Onderzoeksverslag waterstalen – Dialyse – Ringleiding).
    Twee rijen per installatie (begin + einde ringleiding) worden samengevoegd
    tot één staal-dict met 4 meetwaarden.

    Geeft lijst van staal-dicts:
    {
      "datum", "eenheid", "locatie", "maximo",
      "kiemgetal_begin", "endotoxine_begin",
      "kiemgetal_einde", "endotoxine_einde",
      "wo_id": None, "status": None
    }
    """
    import openpyxl
    from datetime import datetime, timedelta
    from collections import OrderedDict

    wb = openpyxl.load_workbook(pad, data_only=True)
    sheet = wb.worksheets[0]
    for ws in wb.worksheets:
        if "water" in ws.title.lower() or "report" in ws.title.lower():
            sheet = ws
            break

    # Zoek headerrij op 'Datum staalname'
    header_row = None
    header_idx = {}
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for j, cell in enumerate(row):
            if isinstance(cell, str) and "datum staalname" in cell.lower():
                header_row = i
                for k, h in enumerate(row):
                    if h is not None:
                        header_idx[str(h).strip().lower()] = k
                break
        if header_row:
            break

    if header_row is None:
        raise ValueError("Headerrij 'Datum staalname' niet gevonden in RO Excel")

    def vind_col(keywords):
        """Geeft kolomindex terug van eerste kolom waarvan de header alle keywords bevat."""
        for key, idx in header_idx.items():
            if all(k.lower() in key for k in keywords):
                return idx
        # Fallback: één keyword volstaat
        for key, idx in header_idx.items():
            if any(k.lower() in key for k in keywords):
                return idx
        return None

    ci_datum   = vind_col(["datum staalname"])
    ci_eenheid = vind_col(["eenheid staalname"])
    ci_lokaal  = vind_col(["lokaal staalname"])
    ci_detail  = vind_col(["staalname detail"])
    ci_maximo  = vind_col(["maximo"])
    ci_kiem    = vind_col(["kiemgetal"])       # 'Totaal kiemgetal CFU/ml'
    ci_endo    = vind_col(["endotoxine"])      # 'Meting endotoxine IU/ml'

    def cel(row, idx):
        """Lees celwaarde veilig als string; integer 0 wordt '0' niet ''."""
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    def parse_datum(datum_raw) -> str:
        try:
            if isinstance(datum_raw, datetime):
                return datum_raw.strftime("%d/%m/%Y")
            if isinstance(datum_raw, (int, float)):
                return (datetime(1899, 12, 30) + timedelta(days=int(datum_raw))).strftime("%d/%m/%Y")
            if isinstance(datum_raw, str) and datum_raw.strip():
                return datum_raw.strip()
        except Exception:
            pass
        return str(datum_raw).strip() if datum_raw is not None else ""

    # Verzamel rijen; bewaar originele volgorde via OrderedDict
    groepen: OrderedDict = OrderedDict()   # (maximo, datum) -> {"begin": dict|None, "einde": dict|None}

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        maximo = cel(row, ci_maximo)
        if not maximo or maximo.lower() in ("nan", "none"):
            continue

        datum_str = parse_datum(row[ci_datum] if ci_datum is not None else None)
        detail    = cel(row, ci_detail).lower()
        is_begin  = "begin" in detail
        is_einde  = "einde" in detail or "eind" in detail

        rij = {
            "datum":      datum_str,
            "eenheid":    cel(row, ci_eenheid),
            "locatie":    cel(row, ci_lokaal),
            "maximo":     maximo,
            "kiemgetal":  cel(row, ci_kiem) or "0",
            "endotoxine": cel(row, ci_endo) or "<0,25",   # RO norm = <0,25 IU/ml
        }

        key = (maximo, datum_str)
        if key not in groepen:
            groepen[key] = {"begin": None, "einde": None}

        if is_begin and not is_einde:
            groepen[key]["begin"] = rij
        elif is_einde and not is_begin:
            groepen[key]["einde"] = rij
        else:
            # Onherkenbaar detail → eerste = begin, tweede = einde
            if groepen[key]["begin"] is None:
                groepen[key]["begin"] = rij
            elif groepen[key]["einde"] is None:
                groepen[key]["einde"] = rij
            # Derde rij voor zelfde maximo+datum: negeren (onverwacht)

    stalen = []
    for (maximo, datum), g in groepen.items():
        b = g["begin"]
        e = g["einde"]

        if b is None and e is None:
            continue  # lege groep, overslaan

        # Als één van beide ontbreekt: gebruik de aanwezige als fallback
        ref = b or e

        stalen.append({
            "datum":            datum,
            "eenheid":          ref["eenheid"],
            "locatie":          ref["locatie"],
            "maximo":           maximo,
            "kiemgetal_begin":  b["kiemgetal"]  if b else "0",
            "endotoxine_begin": b["endotoxine"] if b else "<0,25",
            "kiemgetal_einde":  e["kiemgetal"]  if e else "0",
            "endotoxine_einde": e["endotoxine"] if e else "<0,25",
            "wo_id":            None,
            "status":           None,
        })

    return stalen


async def afwerk_wo_voor_ro_staal(page, staal: dict, stap_log=None) -> dict:
    """
    Vul 4 taaklijen in voor een RO ringleiding werkorder:
      Lijn 0 — Endotoxine begin ringleiding   (taak 1949)
      Lijn 1 — Endotoxine einde ringleiding   (taak 1950)
      Lijn 2 — Microbiologie begin ringleiding (taak 1951)
      Lijn 3 — Microbiologie einde ringleiding (taak 1952)
    """

    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] RO-afwerk {staal.get('maximo','?')} WO{staal.get('wo_id','?')}: {msg}")
        if stap_log:
            stap_log(msg)

    maximo          = staal.get("maximo", "").strip()
    wo_id           = staal.get("wo_id", "").strip().zfill(10)
    datum_str       = staal.get("datum", "")
    eenheid         = staal.get("eenheid", "")
    locatie         = staal.get("locatie", "")
    kiem_begin      = str(staal.get("kiemgetal_begin", "0")).strip()
    endo_begin      = staal.get("endotoxine_begin", "<0,125").strip()
    kiem_einde      = str(staal.get("kiemgetal_einde", "0")).strip()
    endo_einde      = staal.get("endotoxine_einde", "<0,125").strip()

    if not wo_id or wo_id == "0000000000" or wo_id.lower() == "nvt":
        log("⚠ Geen geldig WO-nummer — overslaan")
        return {"ok": False, "wo_id": wo_id, "status": None,
                "fout": "Geen geldig WO-nummer", "maximo": maximo}

    wo_url = (
        PS_PSC_URL
        + "EMPLOYEE/ERP/c/UZFM_MENU.UZFM_WERKORDER.GBL"
        f"?BUSINESS_UNIT=POUZL&UZ_WO_ID={wo_id}"
        "&PAGE=UZFM_WO_ALG&PortalContentProvider=ERP"
        "&PortalCRefLabel=Werkorder&PortalRegistryName=EMPLOYEE"
        "&PortalHostNode=ERP&NoCrumbs=yes&PortalKeyStruct=yes"
    )

    log(f"Stap 1: Navigeren naar WO {wo_id}")
    try:
        await page.goto(wo_url, wait_until="domcontentloaded", timeout=30_000)
        await page.wait_for_selector("#UZFM_WO_UZ_WO_ID", state="visible", timeout=20_000)

        pagina_wo = (await page.locator("#UZFM_WO_UZ_WO_ID").inner_text(timeout=5_000)).strip()
        if wo_id not in pagina_wo and pagina_wo not in wo_id:
            log(f"Stap 1: ❌ Verkeerde WO geopend (verwacht {wo_id}, got {pagina_wo})")
            return {"ok": False, "wo_id": wo_id, "status": None,
                    "fout": f"Verkeerde WO: {pagina_wo}", "maximo": maximo}
        log("Stap 1: ✓ Correcte WO geopend")

        # Rapportage-tab
        log("Stap 2: [KLIK] Rapportage-tab")
        for selector in ["#ICPanel2", "#ICTAB_2"]:
            try:
                btn = await page.query_selector(selector)
                if btn:
                    await page.click(selector)
                    break
            except Exception:
                pass
        await page.wait_for_selector("#UZFM_WO_WRK_UZ_WO_COMMENT", state="visible", timeout=20_000)
        log("Stap 2: ✓ Rapportage-tab geladen")

        # Status → UITGEV
        log("Stap 3: Status instellen op UITGEV")
        await page.select_option("#UZFM_WO_WRK_UZ_WO_STAT_WZ_TECH\\$70\\$", "UITGEV")
        await asyncio.sleep(0.5)

        # Controleer of lijn 0 al ingevuld is
        eerste_taak = ""
        try:
            eerste_taak = (await page.locator("#UZFM_WO_TAAK_UZ_TAAK_ID\\$0").inner_text(timeout=3_000)).strip()
        except Exception:
            pass
        if len(eerste_taak) >= 2:
            log(f"Stap 4: ℹ️ Taaklijn 0 al ingevuld ({eerste_taak}) — WO overslaan")
            return {"ok": True, "wo_id": wo_id, "status": "reeds_ingevuld",
                    "fout": None, "maximo": maximo,
                    "melding": "Reeds ingevuld, overgeslaan"}

        # Commentaar WO — max 60 tekens (PS-veldlimiet)
        # Als LM of M aanwezig: expliciet vermelden welke meting niet telbaar was
        lm_parts = []
        if kiem_begin.strip().upper() in ("LM", "M"):
            lm_parts.append("begin")
        if kiem_einde.strip().upper() in ("LM", "M"):
            lm_parts.append("einde")

        if lm_parts:
            commentaar_wo = f"RO {datum_str} {maximo} ! verhogingen {'+'.join(lm_parts)}"
        else:
            commentaar_wo = f"RO {datum_str} {locatie} ({maximo})"

        commentaar_wo = commentaar_wo[:60]  # PS-veldlimiet
        log(f"Stap 4: Commentaar ← '{commentaar_wo}' ({len(commentaar_wo)} tekens)")
        await page.fill("#UZFM_WO_WRK_UZ_WO_COMMENT", commentaar_wo)

        # Helper: taaklijn invullen
        async def vul_taaklijn(lijn_nr: int, taak_id: str, is_ok: bool, label: str,
                               tekstwaarde: str | None = None):
            if lijn_nr > 0:
                prev = lijn_nr - 1
                log(f"  Nieuwe taaklijn toevoegen (lijn {lijn_nr})")
                await page.click(f"#UZFM_WO_TAAK\\$new\\${prev}\\$\\$0")
                await page.wait_for_selector(f"#UZ_CHANGED_BY\\${lijn_nr}", state="visible", timeout=15_000)

            log(f"  Lijn {lijn_nr}: taak={taak_id}, OK={is_ok}  [{label}]")
            await page.type(f"#UZFM_WO_TAAK_UZ_TAAK_ID\\${lijn_nr}", taak_id, delay=80)
            await page.click("#UZFM_WO_WRK_UZ_WO_COMMENT")
            await asyncio.sleep(1)
            for _ in range(5):
                try:
                    tekst = (await page.locator(f"#UZFM_WO_TAAKDVW_UZ_OMSCHR150\\${lijn_nr}").inner_text(timeout=2_000)).strip()
                    if tekst:
                        log(f"  Lijn {lijn_nr}: omschrijving = '{tekst}'")
                        break
                except Exception:
                    pass
                await asyncio.sleep(1)

            if tekstwaarde is not None:
                # Vrije tekstwaarde invullen (bv. kiemgetal)
                log(f"  Lijn {lijn_nr}: tekstwaarde ← '{tekstwaarde}'")
                await page.type(f"#UZFM_WO_WRK_UZ_WAARDE_TEKST\\${lijn_nr}", tekstwaarde, delay=80)
                await asyncio.sleep(0.5)
            elif is_ok:
                await page.click(f"#UZFM_WO_WRK_UZ_WAARDE_VLAG_JA\\${lijn_nr}")
            else:
                await page.click(f"#UZFM_WO_WRK_UZ_WAARDE_VLAG_NEE\\${lijn_nr}")
            await asyncio.sleep(1)

        # Lijn 0 — Endotoxine begin
        await vul_taaklijn(0, TAAK_RO_ENDO_BEGIN, _ro_endo_ok(endo_begin),
                           f"Endo begin: {endo_begin}")
        # Lijn 1 — Endotoxine einde
        await vul_taaklijn(1, TAAK_RO_ENDO_EINDE, _ro_endo_ok(endo_einde),
                           f"Endo einde: {endo_einde}")
        # Lijn 2 — Microbiologie begin (Ja/Nee)
        await vul_taaklijn(2, TAAK_RO_MICRO_BEGIN, _ro_kiemgetal_ok(kiem_begin),
                           f"Micro begin: {kiem_begin}")
        # Lijn 3 — Microbiologie einde (Ja/Nee)
        await vul_taaklijn(3, TAAK_RO_MICRO_EINDE, _ro_kiemgetal_ok(kiem_einde),
                           f"Micro einde: {kiem_einde}")
        # Lijn 4 — Kiemgetal begin (tekstwaarde)
        await vul_taaklijn(4, TAAK_RO_KIEM_BEGIN, True,
                           f"Kiem begin tekst: {kiem_begin}", tekstwaarde=kiem_begin)
        # Lijn 5 — Kiemgetal einde (tekstwaarde)
        await vul_taaklijn(5, TAAK_RO_KIEM_EINDE, True,
                           f"Kiem einde tekst: {kiem_einde}", tekstwaarde=kiem_einde)

        # Opslaan
        log("Stap 9: [KLIK] Opslaan '#ICSave'")
        await page.click('[id="#ICSave"]')
        await asyncio.sleep(3)

        # Verificatie
        log("Stap 10: Heropen WO voor verificatie...")
        await page.goto(wo_url, wait_until="domcontentloaded", timeout=30_000)
        await page.wait_for_selector("#UZFM_WO_UZ_WO_ID", state="visible", timeout=20_000)
        for selector in ["#ICPanel2", "#ICTAB_2"]:
            try:
                btn = await page.query_selector(selector)
                if btn:
                    await page.click(selector)
                    break
            except Exception:
                pass
        await page.wait_for_selector("#UZFM_WO_WRK_UZ_WO_COMMENT", state="visible", timeout=20_000)

        taak0 = taak1 = taak2 = taak3 = ""
        for i, var in enumerate([taak0, taak1, taak2, taak3]):
            try:
                locals()[f"taak{i}"] = (await page.locator(f"#UZFM_WO_TAAK_UZ_TAAK_ID\\${i}").inner_text(timeout=3_000)).strip()
            except Exception:
                pass
        # Herlees via directe variabelen
        taaklist = []
        for i in range(6):
            try:
                t = (await page.locator(f"#UZFM_WO_TAAK_UZ_TAAK_ID\\${i}").inner_text(timeout=3_000)).strip()
            except Exception:
                t = ""
            taaklist.append(t)

        eindstatus = ""
        try:
            eindstatus = (await page.locator("#UZFM_WO_UZ_WO_STATUS").inner_text(timeout=5_000)).strip()
        except Exception:
            pass

        log(f"Stap 10: taken={taaklist}, status={eindstatus}")

        verwacht = [TAAK_RO_ENDO_BEGIN, TAAK_RO_ENDO_EINDE,
                    TAAK_RO_MICRO_BEGIN, TAAK_RO_MICRO_EINDE,
                    TAAK_RO_KIEM_BEGIN, TAAK_RO_KIEM_EINDE]
        verzonden_ok = (taaklist == verwacht and eindstatus == "UITGEV")

        if verzonden_ok:
            log("Stap 10: ✓ RO resultaten succesvol verzonden en geverifieerd")
            melding = "Resultaten succesvol verzonden"
        else:
            log("Stap 10: ⚠ Verificatie niet volledig geslaagd")
            melding = f"taken={taaklist}, status={eindstatus}"

        return {
            "ok": verzonden_ok,
            "wo_id": wo_id,
            "status": eindstatus,
            "fout": None if verzonden_ok else melding,
            "maximo": maximo,
            "melding": melding,
        }

    except Exception as exc:
        import traceback
        tb = traceback.format_exc()
        log(f"❌ Uitzondering: {exc}")
        for r in tb.splitlines():
            log(f"  TRACEBACK: {r}")
        return {"ok": False, "wo_id": wo_id, "status": None,
                "fout": str(exc), "maximo": maximo}


async def zoek_werkorders_ro(stalen: list[dict], stap_log=None) -> list[dict]:
    """Login eenmalig, zoek voor alle RO-stalen de WO op (hergebruikt zoek_wo_voor_staal)."""
    from playwright.async_api import async_playwright

    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] RO-ZOEK: {msg}")
        if stap_log:
            stap_log(msg)

    resultaten = []
    async with async_playwright() as pw:
        log(f"Browser opstarten (headless={PS_HEADLESS})...")
        browser = await pw.chromium.launch(headless=PS_HEADLESS)
        context = await browser.new_context(viewport={"width": 1280, "height": 900}, locale="nl-BE")
        page    = await context.new_page()
        log("Browser gereed ✓")
        try:
            if not await login(page, log):
                log("❌ Login mislukt")
                for s in stalen:
                    resultaten.append({**s, "ok": False, "wo_id": None,
                                       "status": None, "fout": "Login mislukt"})
                return resultaten

            log(f"✓ Ingelogd — {len(stalen)} RO stalen verwerken")
            for idx, staal in enumerate(stalen):
                log(f"══ Staal {idx+1}/{len(stalen)}: {staal.get('maximo','?')} ({staal.get('datum','?')}) ══")

                def s_log(msg, _s=staal):
                    lbl = f"{_s.get('maximo','?')} | {_s.get('datum','?')}: {msg}"
                    ts  = datetime.now().strftime("%H:%M:%S")
                    print(f"[{ts}] {lbl}")
                    if stap_log:
                        stap_log(lbl)

                # Bouw een tijdelijk staal-dict compatibel met zoek_wo_voor_staal
                staal_td = {**staal, "detail": staal.get("locatie", "")}
                res = await zoek_wo_voor_staal(page, staal_td, stap_log=s_log)
                resultaten.append({**staal, **res})
                await asyncio.sleep(1)
        finally:
            log("Browser afsluiten...")
            await browser.close()
            log("Browser afgesloten ✓")

    return resultaten


async def verzend_resultaten_ro(stalen: list[dict], stap_log=None) -> list[dict]:
    """Login eenmalig, werk voor alle RO-stalen (met wo_id != nvt) de WO af."""
    from playwright.async_api import async_playwright

    def log(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] RO-VERZEND: {msg}")
        if stap_log:
            stap_log(msg)

    resultaten = []
    async with async_playwright() as pw:
        log(f"Browser opstarten (headless={PS_HEADLESS})...")
        browser = await pw.chromium.launch(headless=PS_HEADLESS)
        context = await browser.new_context(viewport={"width": 1280, "height": 900}, locale="nl-BE")
        page    = await context.new_page()
        log("Browser gereed ✓")
        try:
            if not await login(page, log):
                log("❌ Login mislukt")
                for s in stalen:
                    resultaten.append({**s, "ok": False, "fout": "Login mislukt"})
                return resultaten

            log(f"✓ Ingelogd — {len(stalen)} RO stalen ontvangen")
            te_verzenden = [s for s in stalen if s.get("wo_id") and s.get("wo_id") != "nvt"]
            overgeslagen = [s for s in stalen if s not in te_verzenden]

            log(f"  {len(te_verzenden)} te verzenden, {len(overgeslagen)} overgeslagen")
            for s in overgeslagen:
                reden = "Geen WO" if not s.get("wo_id") else f"Status = {s.get('status','?')}"
                log(f"  ↷ Overgeslagen: {s.get('maximo','?')} — {reden}")
                resultaten.append({**s, "ok": False, "melding": reden, "fout": reden})

            for idx, staal in enumerate(te_verzenden):
                log(f"══ Staal {idx+1}/{len(te_verzenden)}: {staal.get('maximo','?')} WO={staal.get('wo_id','?')} ══")

                def s_log(msg, _s=staal):
                    lbl = f"{_s.get('maximo','?')} WO{_s.get('wo_id','?')}: {msg}"
                    ts  = datetime.now().strftime("%H:%M:%S")
                    print(f"[{ts}] {lbl}")
                    if stap_log:
                        stap_log(lbl)

                res = await afwerk_wo_voor_ro_staal(page, staal, stap_log=s_log)
                resultaten.append({**staal, **res})
                await asyncio.sleep(1)
        finally:
            log("Browser afsluiten...")
            await browser.close()
            log("Browser afgesloten ✓")

    return resultaten