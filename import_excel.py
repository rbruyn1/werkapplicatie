"""
import_excel.py
==============
Leest het Thuisdialyse-sheet en staaloverzicht uit het .xlsm bestand
en schrijft thuisdialyse.json.

Gebruik:
    python import_excel.py pad/naar/Toestelstalen.xlsm [--output thuisdialyse.json]
"""
import json
import sys
import argparse
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl


def excel_serial_to_date(serial) -> str | None:
    if not serial:
        return None
    try:
        d = datetime(1899, 12, 30) + timedelta(days=int(serial))
        return d.strftime("%Y-%m-%d")
    except Exception:
        return None


def parse_td_nr(id_str: str) -> int | None:
    """Haal TD-nummer uit strings als 'toestel TD57' of 'Toestel TD64'."""
    s = str(id_str).lower().strip()
    if "td" not in s:
        return None
    try:
        return int(s.split("td")[1].strip())
    except (ValueError, IndexError):
        return None


def lees_laatste_staalname(wb) -> dict[int, str]:
    """Geeft {td_nr: 'YYYY-MM-DD'} met de meest recente staalname per toestel."""
    ws = wb["staaloverzicht"]
    laatste: dict[int, datetime] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        datum, id_str = row[0], row[1]
        if not datum or not id_str:
            continue
        td_nr = parse_td_nr(id_str)
        if td_nr is None:
            continue
        if isinstance(datum, datetime):
            d = datum
        elif isinstance(datum, (int, float)):
            d = datetime(1899, 12, 30) + timedelta(days=int(datum))
        else:
            continue
        if td_nr not in laatste or d > laatste[td_nr]:
            laatste[td_nr] = d
    return {k: v.strftime("%Y-%m-%d") for k, v in laatste.items()}


def lees_toestellen(wb, laatste_staalname: dict) -> list[dict]:
    """Leest de actieve toestellen uit het Thuisdialyse-sheet."""
    ws = wb["Thuisdialyse"]
    toestellen = []
    huidige_patient = None

    for row in ws.iter_rows(values_only=True):
        # Kolommen: A(0) B(1) C(2) D(3) E(4) F(5=checkbox) G(6=WO status)
        b, c, d, f, g = row[1], row[2], row[3], row[5] if len(row) > 5 else None, row[6] if len(row) > 6 else None

        if b and c is None and d is None and isinstance(b, str):
            # Patiëntrij: "Naam - EAD: ... - Adres"
            huidige_patient = str(b).split(" - EAD:")[0].strip()

        elif b == "AK98" and c:
            td_nr = int(c)
            maximo = f"T{d}" if d else None
            toestellen.append({
                "td_nr": td_nr,
                "patient": huidige_patient,
                "maximo": maximo,
                "toestel_type": "AK98",
                "datum_laatste_staalname": laatste_staalname.get(td_nr),
                "laatste_status": g if isinstance(g, str) else None,
                "laatste_wo": None,
                "laatste_wo_id": None,
            })

    return toestellen


def importeer(xlsm_pad: str, output_pad: str):
    wb = openpyxl.load_workbook(xlsm_pad, data_only=True)

    if "Thuisdialyse" not in wb.sheetnames:
        print("❌ Sheet 'Thuisdialyse' niet gevonden in het bestand.", file=sys.stderr)
        sys.exit(1)

    laatste_staalname = lees_laatste_staalname(wb) if "staaloverzicht" in wb.sheetnames else {}
    toestellen = lees_toestellen(wb, laatste_staalname)

    # Bestaand JSON laden om geschiedenis te bewaren
    output_path = Path(output_pad)
    bestaand = {"toestellen": [], "geschiedenis": []}
    if output_path.exists():
        with open(output_path, encoding="utf-8") as f:
            bestaand = json.load(f)

    # Merge: bewaar wo-info van bestaande toestellen
    bestaand_map = {t["td_nr"]: t for t in bestaand.get("toestellen", [])}
    for t in toestellen:
        oud = bestaand_map.get(t["td_nr"], {})
        t["laatste_wo"] = oud.get("laatste_wo", t["laatste_wo"])
        t["laatste_wo_id"] = oud.get("laatste_wo_id", t["laatste_wo_id"])
        if not t["laatste_status"]:
            t["laatste_status"] = oud.get("laatste_status")

    data = {
        "toestellen": toestellen,
        "geschiedenis": bestaand.get("geschiedenis", []),
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✅ {len(toestellen)} toestellen geïmporteerd → {output_pad}")
    for t in toestellen:
        print(f"   TD{t['td_nr']:3d}  {t['patient']:<30}  {t['maximo']}  laatste: {t['datum_laatste_staalname']}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importeer thuisdialyse toestellen uit Excel")
    parser.add_argument("xlsm", help="Pad naar het .xlsm bestand")
    parser.add_argument("--output", default="thuisdialyse.json", help="Output JSON pad")
    args = parser.parse_args()
    importeer(args.xlsm, args.output)
